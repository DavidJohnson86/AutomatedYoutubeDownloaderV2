
import os
from openpyxl import load_workbook

class Parser():
    
    def __init__(self, IN_FILE):
        self.wb = load_workbook(filename = IN_FILE)
        self.current_sheet = self.wb['needs_mp3']
        self.max_row = (len(self.current_sheet['A']))
        self.title_name = 'B'
    
    def WriteFile(self):
        for i in range(3, self.max_row+1):
            with open('input.txt', 'a') as files:
                print(self.current_sheet[self.title_name + str(i)].value + '\n')
                files.write(self.current_sheet[self.title_name + str(i)].value + '\n')


if __name__ =="__main__":
    IN_FILE = os.getcwd() + r'\needs_mp3.xlsm'
    parseexcel = Parser(IN_FILE)
    parseexcel.WriteFile()
    